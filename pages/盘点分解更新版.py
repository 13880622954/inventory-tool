import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(page_title="盘存分析表生成", page_icon="📊", layout="wide")
st.title("📦 盘存分析表生成器")

# ---------- 状态 ----------
for key, default in [('processed', False), ('output_files', None), ('error', None)]:
    if key not in st.session_state:
        st.session_state[key] = default

# ---------- 工具函数 ----------
def clean_str(val):
    if pd.isna(val): return ''
    s = str(val).strip()
    if '.' in s:
        try:
            f = float(s)
            if f.is_integer(): s = str(int(f))
        except: pass
    return s

def clean_float(val):
    try: return float(val)
    except: return 0.0

def get_warehouse_desc(code, phys_dict, gift_dict):
    if pd.isna(code) or code == "": return ""
    code = str(code).strip()
    return gift_dict.get(code, "") if code.endswith("6") else phys_dict.get(code, "")

def assign_age_bucket(days, qty):
    buckets = {
        "3个月库龄":0,"4-6个月库龄":0,"7-12个月库龄":0,
        "1-2年库龄":0,"2-3年库龄":0,"3年以上库龄":0,"10年以上库龄":0
    }
    if days <= 90: buckets["3个月库龄"] = qty
    elif days <= 180: buckets["4-6个月库龄"] = qty
    elif days <= 365: buckets["7-12个月库龄"] = qty
    elif days <= 730: buckets["1-2年库龄"] = qty
    elif days <= 1095: buckets["2-3年库龄"] = qty
    elif days <= 3650: buckets["3年以上库龄"] = qty
    else: buckets["10年以上库龄"] = qty
    return pd.Series(buckets)

# ---------- 核心处理 ----------
def process_files(ib00_bytes, location_bytes, age_bytes, template_bytes, age_gift_bytes=None):
    # 1. 读取文件
    df_ib00 = pd.read_excel(io.BytesIO(ib00_bytes))
    df_phys = pd.read_excel(io.BytesIO(location_bytes), sheet_name="实物库位表")
    try: df_gift_loc = pd.read_excel(io.BytesIO(location_bytes), sheet_name="赠品库位表")
    except: df_gift_loc = pd.DataFrame()
    df_age_main = pd.read_excel(io.BytesIO(age_bytes), sheet_name=0)

    # 2. 仓库描述
    df_ib00["存储位置"] = df_ib00["存储位置"].astype(str).apply(clean_str)
    for df in [df_phys, df_gift_loc]:
        if not df.empty:
            df["库位代码"] = df["库位代码"].astype(str).apply(clean_str)
            df["仓库描述"] = df["仓库描述"].astype(str).apply(clean_str)
    phys_dict = dict(zip(df_phys["库位代码"], df_phys["仓库描述"]))
    gift_dict = dict(zip(df_gift_loc["库位代码"], df_gift_loc["仓库描述"])) if not df_gift_loc.empty else {}
    df_ib00["仓库描述"] = df_ib00["存储位置"].apply(lambda x: get_warehouse_desc(x, phys_dict, gift_dict))

    # 3. 成品库龄
    df_age_main.rename(columns={"物料编码":"物料代码","批号":"产品等级"}, inplace=True)
    for c in ["物料代码","工厂","库位","产品等级","数量","库龄"]:
        if c not in df_age_main.columns: raise KeyError(f"成品库龄表缺列：{c}")
    df_age_main["库位"] = df_age_main["库位"].astype(str).apply(clean_str)
    age_buckets = df_age_main.apply(lambda r: assign_age_bucket(r["库龄"], r["数量"]), axis=1)
    df_age_main = pd.concat([df_age_main, age_buckets], axis=1)
    age_summary_main = df_age_main.groupby(["物料代码","工厂","库位","产品等级"], as_index=False).agg({
        "3个月库龄":"sum","4-6个月库龄":"sum","7-12个月库龄":"sum",
        "1-2年库龄":"sum","2-3年库龄":"sum","3年以上库龄":"sum","10年以上库龄":"sum"
    })

    # 4. 赠品库龄
    age_summary_gift = None
    if age_gift_bytes is not None:
        df_age_gift = pd.read_excel(io.BytesIO(age_gift_bytes), sheet_name=0)
        df_age_gift.columns = [str(c).strip() for c in df_age_gift.columns]
        gift_rename = {"物料":"物料代码","库存地":"库位","批次":"产品等级"}
        df_age_gift.rename(columns=gift_rename, inplace=True)
        for c in ["物料代码","工厂","库位","产品等级"]:
            if c not in df_age_gift.columns: raise KeyError(f"赠品库龄表缺列：{c}")
        df_age_gift["库位"] = df_age_gift["库位"].astype(str).apply(clean_str)

        month_cols = ["1月","2月","3月","4月","5月","6月","7-12月","13-18月","19-24月","2-3年","3年以上"]
        for col in month_cols:
            if col in df_age_gift.columns:
                df_age_gift[col] = pd.to_numeric(df_age_gift[col], errors='coerce').fillna(0)
            else:
                df_age_gift[col] = 0

        df_age_gift["3个月库龄"] = df_age_gift["1月"] + df_age_gift["2月"] + df_age_gift["3月"]
        df_age_gift["4-6个月库龄"] = df_age_gift["4月"] + df_age_gift["5月"] + df_age_gift["6月"]
        df_age_gift["7-12个月库龄"] = df_age_gift["7-12月"]
        df_age_gift["1-2年库龄"] = df_age_gift["13-18月"] + df_age_gift["19-24月"]
        df_age_gift["2-3年库龄"] = df_age_gift["2-3年"]
        df_age_gift["3年以上库龄"] = df_age_gift["3年以上"]
        df_age_gift["10年以上库龄"] = 0

        # 物料代码统一为字符串
        df_age_gift["物料代码"] = df_age_gift["物料代码"].astype(str).str.strip()

        age_summary_gift = df_age_gift.groupby(["物料代码","工厂","库位","产品等级"], as_index=False).agg({
            "3个月库龄":"sum","4-6个月库龄":"sum","7-12个月库龄":"sum",
            "1-2年库龄":"sum","2-3年库龄":"sum","3年以上库龄":"sum","10年以上库龄":"sum"
        })
        age_summary_gift["物料代码"] = age_summary_gift["物料代码"].astype(str).str.strip()

    # 5. 合并
    if "产品等级" not in df_ib00.columns:
        df_ib00["产品等级"] = df_ib00.get("批次", "")
    df_ib00.rename(columns={"存储位置":"库位"}, inplace=True)
    df_ib00["物料代码"] = df_ib00["物料代码"].astype(str).str.strip()
    df_ib00["库位_str"] = df_ib00["库位"].astype(str).apply(clean_str)
    mask_gift = df_ib00["库位_str"].str.endswith("6")
    df_fin = df_ib00[~mask_gift].copy()
    df_gift = df_ib00[mask_gift].copy()

    merge_keys = ["物料代码","工厂","库位","产品等级"]
    df_fin = pd.merge(df_fin, age_summary_main, on=merge_keys, how="left")
    if age_summary_gift is not None:
        df_gift = pd.merge(df_gift, age_summary_gift, on=merge_keys, how="left")
    else:
        df_gift = pd.merge(df_gift, age_summary_main, on=merge_keys, how="left")

    df_merged = pd.concat([df_fin, df_gift], ignore_index=True)
    df_merged.rename(columns={"库位":"存储位置"}, inplace=True)

    age_cols = ["3个月库龄","4-6个月库龄","7-12个月库龄","1-2年库龄","2-3年库龄","3年以上库龄","10年以上库龄"]
    for c in age_cols:
        df_merged[c] = df_merged[c].fillna(0).astype(int)

    df_merged["ERP账面数量"] = df_merged["非限制使用的库存"].apply(clean_float) + df_merged["冻结库存"].apply(clean_float)

    # 6. 模板填充
    warehouses = df_merged["仓库描述"].dropna().unique()
    warehouses = [w for w in warehouses if w != ""]
    output_files = {}

    def find_header_row(ws):
        for row in range(1, ws.max_row+1):
            for col in range(1, ws.max_column+1):
                if ws.cell(row=row, column=col).value == "工厂": return row
        return 2

    def get_column_map(ws, header_row):
        col_map = {}
        for col in range(1, ws.max_column+1):
            cell1 = ws.cell(row=header_row, column=col)
            cell2 = ws.cell(row=header_row+1, column=col) if header_row+1 <= ws.max_row else None
            name = None
            if cell2 and cell2.value and isinstance(cell2.value, str):
                name = cell2.value.replace('\n','').strip()
            elif cell1.value and isinstance(cell1.value, str):
                name = cell1.value.replace('\n','').strip()
            if name: col_map[col] = name
        return col_map

    def write_data_rows(ws, data_start_row, data_rows, col_map, header_row, has_total):
        bottom_row = None
        for row in range(data_start_row, ws.max_row+1):
            for col in range(1, ws.max_column+1):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str) and "电子版：" in val:
                    bottom_row = row; break
            if bottom_row: break
        if not bottom_row: bottom_row = ws.max_row+1
        if bottom_row > data_start_row: ws.delete_rows(data_start_row, bottom_row-data_start_row)

        new_col = ws.max_column+1
        ws.cell(row=header_row, column=new_col, value="是否符合")
        if not data_rows: return

        ws.insert_rows(data_start_row, amount=len(data_rows))
        for i, row_cells in enumerate(data_rows):
            for col_num, val in enumerate(row_cells, start=1):
                cell = ws.cell(row=data_start_row+i, column=col_num)
                if cell.__class__.__name__ == 'MergedCell': continue
                if val is not None: cell.value = val

        actual_col = adj_col = diff_col = erp_col = in_col = out_col = None
        age_col_nums = []
        for col_num, col_name in col_map.items():
            if col_name == "实盘数量": actual_col = col_num
            elif col_name == "调整后数量": adj_col = col_num
            elif col_name == "盘盈（+）盘亏（-）数量": diff_col = col_num
            elif col_name == "ERP账面数量": erp_col = col_num
            elif col_name == "入库未记数": in_col = col_num
            elif col_name == "出库未记数": out_col = col_num
            elif col_name in age_cols: age_col_nums.append(col_num)

        rows_count = len(data_rows) - (1 if has_total else 0)
        if rows_count>0 and actual_col and adj_col and diff_col and erp_col and in_col and out_col and age_col_nums:
            age_start = min(age_col_nums); age_end = max(age_col_nums)
            for i in range(rows_count):
                r = data_start_row+i
                cond1 = f"N({get_column_letter(actual_col)}{r})=N({get_column_letter(adj_col)}{r})+N({get_column_letter(diff_col)}{r})"
                cond2 = f"N({get_column_letter(actual_col)}{r})=SUM({get_column_letter(age_start)}{r}:{get_column_letter(age_end)}{r})"
                cond3 = f"N({get_column_letter(adj_col)}{r})=N({get_column_letter(erp_col)}{r})+N({get_column_letter(in_col)}{r})-N({get_column_letter(out_col)}{r})"
                ws.cell(row=r, column=new_col, value=f'=IF(AND({cond1},{cond2},{cond3}),"是","否")')
        if has_total:
            ws.cell(row=data_start_row+len(data_rows)-1, column=new_col, value="")

    now = datetime.now()
    last_month = (now.replace(day=1) - pd.DateOffset(days=1)).strftime("%Y年%m月")

    for warehouse in warehouses:
        df_w = df_merged[df_merged["仓库描述"] == warehouse].copy()
        df_w["存储位置_str"] = df_w["存储位置"].astype(str).apply(clean_str)
        df_fin_w = df_w[~df_w["存储位置_str"].str.endswith("6")]
        df_gift_w = df_w[df_w["存储位置_str"].str.endswith("6")]

        safe_name = warehouse.replace("/","_").replace("\\","_").replace(":","_")
        fname = f"{last_month}合肥美菱集团控股有限公司内（外）销产成品盘存分析表--{safe_name}.xlsx"
        wb = load_workbook(io.BytesIO(template_bytes))
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "请输入库位" in cell.value: cell.value = cell.value.replace("请输入库位", warehouse)
                        if "2026年3月" in cell.value: cell.value = cell.value.replace("2026年3月", last_month)

        for sheet, df_data in [("成品", df_fin_w), ("赠品", df_gift_w)]:
            if sheet not in wb.sheetnames: continue
            ws = wb[sheet]
            hdr = find_header_row(ws)
            start_row = hdr+2
            col_map = get_column_map(ws, hdr)

            rows = []
            total_qty = 0
            if not df_data.empty:
                for _, row in df_data.iterrows():
                    rc = [None]*ws.max_column
                    for cnum, cname in col_map.items():
                        if cname == "工厂": val = row.get("工厂","")
                        elif cname == "库位": val = row.get("存储位置","")
                        elif cname == "库位名称": val = warehouse
                        elif cname == "物料代码": val = row.get("物料代码","")
                        elif cname == "物料描述": val = row.get("物料描述","")
                        elif cname == "产品等级": val = row.get("产品等级","")
                        elif cname == "单位": val = row.get("单位","")
                        elif cname == "ERP账面数量":
                            val = row.get("ERP账面数量",0); total_qty += val
                        elif cname in ["ERP账面金额","入库未记数","出库未记数","调整后数量","实盘数量","盘盈（+）盘亏（-）数量"]:
                            val = ""
                        elif cname in age_cols:
                            val = row.get(cname, 0)
                        else:
                            val = ""
                        rc[cnum-1] = val
                    rows.append(rc)
                total_r = [None]*ws.max_column
                for cnum, cname in col_map.items():
                    if cname == "库位名称": total_r[cnum-1] = "合计"
                    elif cname == "ERP账面数量": total_r[cnum-1] = total_qty
                rows.append(total_r)
                has_total = True
            else:
                has_total = False
            write_data_rows(ws, start_row, rows, col_map, hdr, has_total)

        with io.BytesIO() as buf:
            wb.save(buf)
            output_files[fname] = buf.getvalue()

    # 汇总表
    df_merged["类型"] = df_merged["存储位置"].apply(lambda x: "赠品" if x.endswith("6") else "成品")
    summary = df_merged.groupby(["仓库描述","类型"])["ERP账面数量"].sum().reset_index()
    pivot = summary.pivot(index="仓库描述", columns="类型", values="ERP账面数量").fillna(0).reset_index()
    for t in ["成品","赠品"]:
        if t not in pivot.columns: pivot[t] = 0
    pivot["总计"] = pivot["成品"] + pivot["赠品"]
    pivot = pivot.sort_values("总计", ascending=False)
    total_row = pd.DataFrame([["总计", pivot["成品"].sum(), pivot["赠品"].sum(), pivot["总计"].sum()]],
                             columns=["仓库描述","成品","赠品","总计"])
    pivot = pd.concat([pivot, total_row], ignore_index=True)
    sum_bytes = io.BytesIO()
    pivot.to_excel(sum_bytes, index=False)
    output_files[f"{last_month}库存盘点汇总表.xlsx"] = sum_bytes.getvalue()

    return output_files

# ---------- UI ----------
st.markdown("### 📤 上传文件")
col1, col2 = st.columns(2)
with col1:
    f_ib00 = st.file_uploader("1. IB00库存表", type=['xlsx','xls','csv'], key="ib00")
    f_loc = st.file_uploader("2. 库位表", type=['xlsx','xls','csv'], key="loc")
    f_age = st.file_uploader("3. 成品库龄报表", type=['xlsx','xls','csv'], key="age")
with col2:
    f_tpl = st.file_uploader("4. 模板文件", type=['xlsx','xls'], key="tpl")
    f_age_gift = st.file_uploader("5. 赠品库龄报表（可选）", type=['xlsx','xls','csv'], key="age_gift")

if st.button("🚀 生成盘存分析表", type="primary", use_container_width=True):
    if not all([f_ib00, f_loc, f_age, f_tpl]):
        st.error("❌ 前四个文件为必传项！")
    else:
        with st.spinner("⏳ 处理中..."):
            try:
                outputs = process_files(
                    f_ib00.read(), f_loc.read(), f_age.read(), f_tpl.read(),
                    f_age_gift.read() if f_age_gift else None
                )
                st.session_state.processed = True
                st.session_state.output_files = outputs
                st.session_state.error = None
                st.success("✅ 完成！")
            except Exception as e:
                st.session_state.processed = False
                st.session_state.output_files = None
                st.session_state.error = str(e)
                st.error(f"❌ 错误：{e}")

if st.session_state.processed and st.session_state.output_files:
    st.subheader("📥 下载结果")
    for name, data in st.session_state.output_files.items():
        if "汇总表" in name:
            st.write("📊 汇总表预览")
            st.dataframe(pd.read_excel(io.BytesIO(data)), use_container_width=True)
            break
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for fname, fdata in st.session_state.output_files.items():
            zf.writestr(fname, fdata)
    zip_buffer.seek(0)
    st.download_button("📦 下载全部结果 (ZIP)", data=zip_buffer, file_name="盘存分析表结果.zip", mime="application/zip")

if st.session_state.error:
    st.error(f"⚠️ 上次错误：{st.session_state.error}")
