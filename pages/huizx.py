import streamlit as st
import pandas as pd
import zipfile
import io
import warnings
from openpyxl import load_workbook
from datetime import datetime

warnings.filterwarnings('ignore')

# ========== 配置 ==========
SHEET_PHYSICAL = '实物库位表'
SHEET_GIFT = '赠品库位表'
INVENTORY_SHEET_PRODUCT = '成品'
INVENTORY_SHEET_GIFT = '赠品'
COL_LOCATION_CODE = '库位代码'
COL_LOCATION_DESC = '仓库描述'

FIXED_COLUMNS = [
    '工厂', '库位', '库位名称', '物料代码', '物料描述', '产品等级', '单位',
    'ERP账面数量', 'ERP账面金额', '入库未记数', '出库未记数', '调整后数量',
    '实盘数量', '盘盈（+）', '差异原因分析', '实物状态', '是否影响正常销售',
    '产品账实等级是否一致', '3个月库龄', '4-6个月库龄', '7-12个月库龄',
    '1-2年库龄', '2-3年库龄', '3年以上库龄', '10年以上库龄',
    '计提跌价准备金额', '实物状态是否为裸机'
]

SUM_COLUMNS = [
    'ERP账面数量', '入库未记数', '出库未记数', '调整后数量', '实盘数量', '盘盈（+）'
]

# ========== 初始化 session_state 缓存 ==========
if 'cached_files' not in st.session_state:
    st.session_state.cached_files = []  # 每个元素为 {'name': 文件名, 'data': bytes}

# ========== 辅助函数（与原脚本一致） ==========
def clean_str(val):
    if pd.isna(val):
        return ''
    return str(val).strip()

def extract_location_dict_from_bytes(file_bytes, sheet_name):
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name)
        if COL_LOCATION_CODE not in df.columns:
            st.error(f"{sheet_name} 中缺少列: {COL_LOCATION_CODE}")
            return {}
        desc_col = COL_LOCATION_DESC if COL_LOCATION_DESC in df.columns else COL_LOCATION_CODE
        location_dict = {}
        for _, row in df.iterrows():
            code = clean_str(row[COL_LOCATION_CODE])
            desc = clean_str(row[desc_col]) if desc_col in df.columns else code
            if code:
                location_dict[code] = desc
        return location_dict
    except Exception as e:
        st.error(f"读取 {sheet_name} 失败: {e}")
        return {}

def find_two_row_header(df):
    header_bottom_idx = None
    for idx in range(min(50, len(df))):
        row = df.iloc[idx]
        row_str = ' '.join([clean_str(v) for v in row.values if pd.notna(v)])
        if ('库位' in row_str) and ('物料代码' in row_str or '物料描述' in row_str or '工厂' in row_str):
            header_bottom_idx = idx
            break
    if header_bottom_idx is None:
        return None, None, None
    header_top_idx = header_bottom_idx - 1
    if header_top_idx < 0:
        header_top_idx = None
    data_start = None
    for j in range(header_bottom_idx + 1, len(df)):
        first_cell = clean_str(df.iloc[j, 0]) if df.shape[1] > 0 else ''
        if first_cell and '合计' not in first_cell:
            data_start = j
            break
    return header_top_idx, header_bottom_idx, data_start

def combine_two_row_header(df, header_top_idx, header_bottom_idx):
    top_row = [clean_str(v) for v in df.iloc[header_top_idx].values] if header_top_idx is not None else []
    bottom_row = [clean_str(v) for v in df.iloc[header_bottom_idx].values]
    max_len = max(len(top_row), len(bottom_row))
    top_row += [''] * (max_len - len(top_row))
    bottom_row += [''] * (max_len - len(bottom_row))
    last_non_empty = ''
    for i in range(max_len):
        if top_row[i]:
            last_non_empty = top_row[i]
        else:
            top_row[i] = last_non_empty
    combined = []
    for i in range(max_len):
        top = top_row[i]
        bottom = bottom_row[i]
        if bottom:
            combined.append(f"{top}\n{bottom}" if top else bottom)
        else:
            combined.append(top)
    seen = {}
    unique = []
    for col in combined:
        if col in seen:
            seen[col] += 1
            new_col = f"{col}_{seen[col]}"
        else:
            seen[col] = 0
            new_col = col
        unique.append(new_col)
    return unique

def extract_matched_rows_from_bytes(file_bytes, sheet_name, location_dict):
    try:
        df = pd.read_excel(io.BytesIO(file_bytes), sheet_name=sheet_name, header=None)
    except Exception:
        return pd.DataFrame()
    if df is None or df.empty:
        return pd.DataFrame()
    header_top, header_bottom, data_start = find_two_row_header(df)
    if header_bottom is None or data_start is None:
        return pd.DataFrame()
    headers = combine_two_row_header(df, header_top, header_bottom)
    data_rows = []
    for idx in range(data_start, len(df)):
        row = df.iloc[idx]
        first_cell = clean_str(row.iloc[0]) if len(row) > 0 else ''
        if first_cell == '' or '合计' in first_cell:
            break
        data_rows.append(row.values)
    if not data_rows:
        return pd.DataFrame()
    num_cols = len(data_rows[0])
    temp_columns = [f'col_{i}' for i in range(num_cols)]
    df_data = pd.DataFrame(data_rows, columns=temp_columns)
    location_col_idx = None
    for idx, col_name in enumerate(headers):
        if '库位' in col_name and '库位名称' not in col_name:
            location_col_idx = idx
            break
    if location_col_idx is None:
        return pd.DataFrame()
    location_col = f'col_{location_col_idx}'
    df_data['库位代码'] = df_data[location_col].astype(str).str.strip()
    df_data['仓库描述'] = df_data['库位代码'].map(location_dict).fillna('')
    matched = df_data[df_data['仓库描述'] != ''].copy()
    matched.drop('库位代码', axis=1, inplace=True)
    return matched

def process_inventory_zip(zip_bytes, product_location_dict, gift_location_dict):
    all_product = []
    all_gift = []
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as z:
        for file_name in z.namelist():
            if file_name.endswith('/') or file_name.startswith('~$'):
                continue
            if not (file_name.endswith('.xlsx') or file_name.endswith('.xls')):
                continue
            try:
                with z.open(file_name) as f:
                    file_bytes = f.read()
                product_data = extract_matched_rows_from_bytes(file_bytes, INVENTORY_SHEET_PRODUCT, product_location_dict)
                if not product_data.empty:
                    all_product.append(product_data)
                gift_data = extract_matched_rows_from_bytes(file_bytes, INVENTORY_SHEET_GIFT, gift_location_dict)
                if not gift_data.empty:
                    all_gift.append(gift_data)
            except Exception as e:
                st.warning(f"处理文件 {file_name} 出错: {e}")
    combined_product = pd.concat(all_product, ignore_index=True) if all_product else pd.DataFrame()
    combined_gift = pd.concat(all_gift, ignore_index=True) if all_gift else pd.DataFrame()
    return combined_product, combined_gift

def align_to_fixed_columns(df, fixed_cols):
    if df.empty:
        return pd.DataFrame(columns=fixed_cols)
    num_data_cols = len(df.columns)
    num_fixed = len(fixed_cols)
    result = pd.DataFrame(index=df.index, columns=fixed_cols)
    for i in range(min(num_data_cols, num_fixed)):
        result.iloc[:, i] = df.iloc[:, i]
    return result

def summarize_by_warehouse(df):
    if df.empty:
        return pd.DataFrame(columns=['仓库描述', 'ERP账面数_汇总', '入库未计数', '出库未记数', '调整后数量', '实盘', '盘盈', '盘亏'])
    if '仓库描述' not in df.columns:
        return pd.DataFrame()
    sum_indices = []
    for col in SUM_COLUMNS:
        if col in FIXED_COLUMNS:
            idx = FIXED_COLUMNS.index(col)
            if idx < len(df.columns):
                sum_indices.append(idx)
    for idx in sum_indices:
        col_name = f'col_{idx}'
        df[col_name] = pd.to_numeric(df.iloc[:, idx], errors='coerce').fillna(0)
    grouped = df.groupby('仓库描述')
    agg_dict = {f'col_{idx}': 'sum' for idx in sum_indices}
    result = grouped.agg(agg_dict).reset_index()
    result.rename(columns={'仓库描述': '仓库描述'}, inplace=True)
    rename_map = {}
    for orig, target in zip([f'col_{idx}' for idx in sum_indices], SUM_COLUMNS):
        rename_map[orig] = target
    result.rename(columns=rename_map, inplace=True)
    if '盘盈（+）' in result.columns:
        result['盘盈'] = result['盘盈（+）'].apply(lambda x: x if x > 0 else 0)
        result['盘亏'] = result['盘盈（+）'].apply(lambda x: abs(x) if x < 0 else 0)
        result.drop('盘盈（+）', axis=1, inplace=True)
    else:
        result['盘盈'] = 0
        result['盘亏'] = 0
    final_rename = {
        'ERP账面数量': 'ERP账面数_汇总',
        '入库未记数': '入库未计数',
        '出库未记数': '出库未记数',
        '调整后数量': '调整后数量',
        '实盘数量': '实盘'
    }
    for old, new in final_rename.items():
        if old in result.columns:
            result.rename(columns={old: new}, inplace=True)
    final_cols = ['仓库描述', 'ERP账面数_汇总', '入库未计数', '出库未记数', '调整后数量', '实盘', '盘盈', '盘亏']
    for col in final_cols:
        if col not in result.columns:
            result[col] = 0
    return result[final_cols]

def update_match_file(match_file_bytes, product_summary, gift_summary):
    """使用 openpyxl 直接修改匹配文件，保留所有格式"""
    if match_file_bytes is None:
        return None
    try:
        wb = load_workbook(io.BytesIO(match_file_bytes))
    except Exception as e:
        st.error(f"无法加载匹配文件: {e}")
        return None

    product_sheet_name = None
    gift_sheet_name = None
    for sheet in wb.sheetnames:
        if '成品' in sheet:
            product_sheet_name = sheet
        elif '赠品' in sheet:
            gift_sheet_name = sheet

    if product_sheet_name and not product_summary.empty:
        ws = wb[product_sheet_name]
        header_row = 2
        warehouse_col = None
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value
            if cell_val and '仓库描述' in str(cell_val):
                warehouse_col = col
                break
        if warehouse_col:
            summary_dict = {row['仓库描述']: row for _, row in product_summary.iterrows()}
            for row in range(3, ws.max_row + 1):
                warehouse = ws.cell(row=row, column=warehouse_col).value
                if warehouse and warehouse in summary_dict:
                    new_data = summary_dict[warehouse]
                    col_updates = {
                        'ERP账面数_汇总': 'ERP账面数-仓库数量',
                        '入库未计数': '入库未计数',
                        '出库未记数': '出库未记数',
                        '实盘': '实盘',
                        '盘盈': '盘盈',
                        '盘亏': '盘亏'
                    }
                    for sum_field, target_col in col_updates.items():
                        target_col_idx = None
                        for c in range(1, ws.max_column + 1):
                            if ws.cell(row=header_row, column=c).value == target_col:
                                target_col_idx = c
                                break
                        if target_col_idx:
                            ws.cell(row=row, column=target_col_idx, value=new_data[sum_field])
        else:
            st.warning(f"成品 sheet '{product_sheet_name}' 中未找到仓库描述列，跳过更新")

    if gift_sheet_name and not gift_summary.empty:
        ws = wb[gift_sheet_name]
        header_row = 2
        warehouse_col = None
        for col in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=header_row, column=col).value
            if cell_val and '仓库描述' in str(cell_val):
                warehouse_col = col
                break
        if warehouse_col:
            summary_dict = {row['仓库描述']: row for _, row in gift_summary.iterrows()}
            for row in range(3, ws.max_row + 1):
                warehouse = ws.cell(row=row, column=warehouse_col).value
                if warehouse and warehouse in summary_dict:
                    new_data = summary_dict[warehouse]
                    col_updates = {
                        'ERP账面数_汇总': 'ERP账面数-仓库数量',
                        '入库未计数': '入库未计数',
                        '出库未记数': '出库未记数',
                        '实盘': '实盘',
                        '盘盈': '盘盈',
                        '盘亏': '盘亏'
                    }
                    for sum_field, target_col in col_updates.items():
                        target_col_idx = None
                        for c in range(1, ws.max_column + 1):
                            if ws.cell(row=header_row, column=c).value == target_col:
                                target_col_idx = c
                                break
                        if target_col_idx:
                            ws.cell(row=row, column=target_col_idx, value=new_data[sum_field])
        else:
            st.warning(f"赠品 sheet '{gift_sheet_name}' 中未找到仓库描述列，跳过更新")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output

# ========== 缓存管理函数 ==========
def add_to_cache(file_bytes, original_filename):
    """将更新后的文件添加到缓存"""
    # 生成唯一文件名（原文件名 + 时间戳）
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = original_filename.rsplit('.', 1)[0]
    new_name = f"{base_name}_更新_{timestamp}.xlsx"
    st.session_state.cached_files.append({
        'name': new_name,
        'data': file_bytes
    })
    st.success(f"已缓存: {new_name} (当前共 {len(st.session_state.cached_files)} 个文件)")

def clear_cache():
    st.session_state.cached_files = []
    st.success("缓存已清空")

def download_all_as_zip():
    """将所有缓存文件打包成 ZIP 并返回字节流"""
    if not st.session_state.cached_files:
        return None
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
        for item in st.session_state.cached_files:
            zf.writestr(item['name'], item['data'])
    zip_buffer.seek(0)
    return zip_buffer

# ========== Streamlit 界面 ==========
st.set_page_config(page_title="盘点表汇总工具", layout="wide")
st.title("📊 盘点表批量汇总工具（支持多文件缓存）")

st.markdown("""
**使用说明**：
1. 上传库位表（Excel，包含“实物库位表”和“赠品库位表”两个sheet）。
2. 上传盘点表压缩包（ZIP，内含多个盘点表Excel文件）。
3. 可选上传匹配文件（例如“2026年2月美菱IB00工厂盘存数据、账外物资汇总.xlsx”）。
4. 点击“开始处理”，会生成更新后的匹配文件，并提供“立即下载”和“加入缓存”两个选项。
5. 可以重复处理不同月份的匹配文件，每次可将结果加入缓存。
6. 最后点击“打包下载所有缓存文件”，一次性下载所有缓存的结果。
""")

# 侧边栏上传
with st.sidebar:
    st.header("1. 上传必需文件")
    location_file = st.file_uploader("库位表 (Excel)", type=['xlsx'])
    inventory_zip = st.file_uploader("盘点表压缩包 (ZIP)", type=['zip'])
    st.header("2. 可选匹配文件")
    match_file = st.file_uploader("匹配文件 (Excel，月份可变)", type=['xlsx'])

    process_btn = st.button("开始处理")

# 显示缓存区
st.sidebar.header("📦 缓存区")
if st.sidebar.button("清空缓存"):
    clear_cache()
if st.sidebar.button("📥 打包下载全部缓存"):
    zip_data = download_all_as_zip()
    if zip_data:
        st.sidebar.download_button(
            "点击下载所有缓存文件 (ZIP)",
            zip_data,
            file_name="所有更新文件.zip",
            mime="application/zip"
        )
    else:
        st.sidebar.info("暂无缓存文件")

if st.session_state.cached_files:
    st.sidebar.write(f"已缓存 {len(st.session_state.cached_files)} 个文件：")
    for item in st.session_state.cached_files:
        st.sidebar.text(f"📄 {item['name']}")
else:
    st.sidebar.info("暂无缓存")

# 主处理逻辑
if process_btn:
    if not location_file or not inventory_zip:
        st.error("请同时上传库位表和盘点表压缩包")
        st.stop()

    with st.spinner("正在处理，请稍候..."):
        # 1. 读取库位表
        location_bytes = location_file.read()
        product_location_dict = extract_location_dict_from_bytes(location_bytes, SHEET_PHYSICAL)
        gift_location_dict = extract_location_dict_from_bytes(location_bytes, SHEET_GIFT)
        if not product_location_dict:
            st.error("实物库位表为空或格式错误")
            st.stop()
        st.success(f"实物库位映射: {len(product_location_dict)} 个")
        st.success(f"赠品库位映射: {len(gift_location_dict)} 个")

        # 2. 处理盘点表ZIP
        product_detail, gift_detail = process_inventory_zip(
            inventory_zip.getvalue(), product_location_dict, gift_location_dict
        )
        st.write(f"从盘点表提取的成品明细行数: {len(product_detail)}")
        st.write(f"从盘点表提取的赠品明细行数: {len(gift_detail)}")

        if product_detail.empty and gift_detail.empty:
            st.error("没有匹配到任何数据，请检查库位表和盘点表文件")
            st.stop()

        # 3. 生成汇总和明细
        product_summary = summarize_by_warehouse(product_detail.copy())
        gift_summary = summarize_by_warehouse(gift_detail.copy())
        product_output = align_to_fixed_columns(product_detail, FIXED_COLUMNS)
        gift_output = align_to_fixed_columns(gift_detail, FIXED_COLUMNS)

        # 4. 显示结果
        st.subheader("汇总结果")
        col1, col2 = st.columns(2)
        with col1:
            st.write("**成品按仓库汇总**")
            st.dataframe(product_summary)
        with col2:
            st.write("**赠品按仓库汇总**")
            st.dataframe(gift_summary)

        with st.expander("查看明细"):
            if not product_output.empty:
                st.write("**成品明细**")
                st.dataframe(product_output)
            if not gift_output.empty:
                st.write("**赠品明细**")
                st.dataframe(gift_output)

        # 5. 下载按钮（明细和汇总的CSV）
        st.subheader("下载明细/汇总（CSV）")
        if not product_summary.empty:
            st.download_button("下载成品汇总 (CSV)", product_summary.to_csv(index=False).encode('utf-8-sig'), "成品汇总.csv", "text/csv")
        if not gift_summary.empty:
            st.download_button("下载赠品汇总 (CSV)", gift_summary.to_csv(index=False).encode('utf-8-sig'), "赠品汇总.csv", "text/csv")
        if not product_output.empty:
            st.download_button("下载成品明细 (CSV)", product_output.to_csv(index=False).encode('utf-8-sig'), "成品明细.csv", "text/csv")
        if not gift_output.empty:
            st.download_button("下载赠品明细 (CSV)", gift_output.to_csv(index=False).encode('utf-8-sig'), "赠品明细.csv", "text/csv")

        # 6. 如果上传了匹配文件，则更新并提供下载选项
        if match_file is not None:
            st.info("正在根据新汇总数据更新匹配文件...")
            updated_file = update_match_file(match_file.getvalue(), product_summary, gift_summary)
            if updated_file:
                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        "📥 立即下载更新后的匹配文件",
                        updated_file,
                        file_name=f"更新_{match_file.name}",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                with col2:
                    if st.button("➕ 加入缓存"):
                        add_to_cache(updated_file, match_file.name)
            else:
                st.warning("匹配文件更新失败，请检查文件格式")
        else:
            st.info("未上传匹配文件，如需匹配请上传并重新处理")

    st.success("本次处理完成！")