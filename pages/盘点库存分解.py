import streamlit as st
import pandas as pd
import io
import zipfile
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime

st.set_page_config(page_title="盘存分析表生成", layout="wide")
st.title("盘存分析表生成器")

with st.expander("使用说明"):
    st.markdown("""
    1. **IB00库存表**：必须包含列 `存储位置`、`工厂`、`物料代码`、`物料描述`、`单位`、`非限制使用的库存`、`冻结库存` 等。
    2. **库位表**：必须包含工作表 `实物库位表` 和 `赠品库位表`，各有 `库位代码`、`仓库描述` 列。
    3. **库龄报表**：必须包含列 `物料编码`、`工厂`、`库位`、`批号`、`数量`、`库龄`。
    4. **模板文件**：须有工作表 `成品` 和 `赠品`，表头包含 `工厂`、`库位`、`库位名称`、`物料代码`、`物料描述`、`产品等级`、`单位`、`ERP账面数量`、`ERP账面金额`、`入库未记数`、`出库未记数`、`调整后数量`、`实盘数量`、`盘盈（+）盘亏（-）数量`、库龄列等。
    5. 点击生成后，将自动处理并下载 ZIP 压缩包，内含各仓库的盘存分析表和汇总表。
    """)

# 上传文件
col1, col2 = st.columns(2)
with col1:
    ib00_file = st.file_uploader("1. 上传 IB00 库存表 (.xlsx)", type="xlsx")
    location_file = st.file_uploader("2. 上传库位表 (.xlsx)", type="xlsx")
with col2:
    age_file = st.file_uploader("3. 上传库龄报表 (.xlsx)", type="xlsx")
    template_file = st.file_uploader("4. 上传模板文件 (.xlsx)", type="xlsx")

def process_files(ib00_file, location_file, age_file, template_file):
    """核心处理逻辑，返回 {文件名: 字节} 字典"""
    # 读取四个文件
    df_ib00 = pd.read_excel(ib00_file)
    df_physical = pd.read_excel(location_file, sheet_name="实物库位表")
    df_gift = pd.read_excel(location_file, sheet_name="赠品库位表")
    df_age = pd.read_excel(age_file, sheet_name=0)
    template_wb = load_workbook(template_file)

    # ---------- 1. 匹配仓库描述 ----------
    df_ib00["存储位置_clean"] = df_ib00["存储位置"].astype(str).str.strip()
    df_physical["库位代码_clean"] = df_physical["库位代码"].astype(str).str.strip()
    df_gift["库位代码_clean"] = df_gift["库位代码"].astype(str).str.strip()
    physical_dict = dict(zip(df_physical["库位代码_clean"], df_physical["仓库描述"]))
    gift_dict = dict(zip(df_gift["库位代码_clean"], df_gift["仓库描述"]))

    def get_warehouse_desc(storage_code):
        if pd.isna(storage_code) or storage_code == "":
            return ""
        storage_code = str(storage_code).strip()
        if storage_code.endswith("6"):
            return gift_dict.get(storage_code, "")
        else:
            return physical_dict.get(storage_code, "")

    df_ib00["仓库描述"] = df_ib00["存储位置_clean"].apply(get_warehouse_desc)
    col_index = df_ib00.columns.get_loc("存储位置")
    warehouse_desc_col = df_ib00.pop("仓库描述")
    df_ib00.insert(col_index + 1, "仓库描述", warehouse_desc_col)
    df_ib00.drop(columns=["存储位置_clean"], inplace=True)

    # ---------- 2. 处理库龄报表 ----------
    rename_map = {"物料编码": "物料代码", "批号": "产品等级"}
    df_age.rename(columns=rename_map, inplace=True)
    required_cols = ["物料代码", "工厂", "库位", "产品等级", "数量", "库龄"]
    for col in required_cols:
        if col not in df_age.columns:
            raise KeyError(f"库龄报表中缺少列：{col}")

    df_age["存储位置_clean"] = df_age["库位"].astype(str).str.strip()
    df_age["仓库描述"] = df_age["存储位置_clean"].apply(get_warehouse_desc)
    df_age.drop(columns=["存储位置_clean"], inplace=True)

    def assign_age_bucket(days, qty):
        buckets = {
            "3个月库龄": 0, "4-6个月库龄": 0, "7-12个月库龄": 0,
            "1-2年库龄": 0, "2-3年库龄": 0, "3年以上库龄": 0, "10年以上库龄": 0
        }
        if days <= 90:
            buckets["3个月库龄"] = qty
        elif days <= 180:
            buckets["4-6个月库龄"] = qty
        elif days <= 365:
            buckets["7-12个月库龄"] = qty
        elif days <= 730:
            buckets["1-2年库龄"] = qty
        elif days <= 1095:
            buckets["2-3年库龄"] = qty
        elif days <= 3650:
            buckets["3年以上库龄"] = qty
        else:
            buckets["10年以上库龄"] = qty
        return pd.Series(buckets)

    age_buckets = df_age.apply(lambda row: assign_age_bucket(row["库龄"], row["数量"]), axis=1)
    df_age = pd.concat([df_age, age_buckets], axis=1)
    group_cols = ["物料代码", "工厂", "库位", "产品等级"]
    age_summary = df_age.groupby(group_cols, as_index=False).agg({
        "3个月库龄": "sum", "4-6个月库龄": "sum", "7-12个月库龄": "sum",
        "1-2年库龄": "sum", "2-3年库龄": "sum", "3年以上库龄": "sum", "10年以上库龄": "sum"
    })

    # ---------- 3. 合并库龄数据 ----------
    if "产品等级" not in df_ib00.columns:
        df_ib00["产品等级"] = df_ib00.get("批次", "")
    df_ib00.rename(columns={"存储位置": "库位"}, inplace=True)
    merge_keys = ["物料代码", "工厂", "库位", "产品等级"]
    df_merged = pd.merge(df_ib00, age_summary, on=merge_keys, how="left")
    df_merged.rename(columns={"库位": "存储位置"}, inplace=True)
    age_cols = ["3个月库龄", "4-6个月库龄", "7-12个月库龄", "1-2年库龄", "2-3年库龄", "3年以上库龄", "10年以上库龄"]
    for col in age_cols:
        df_merged[col] = df_merged[col].fillna(0).astype(int)

    # ---------- 4. 生成盘存分析表 ----------
    warehouses = df_merged["仓库描述"].dropna().unique()
    warehouses = [w for w in warehouses if w != ""]
    output_files = {}

    def find_header_row(ws):
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=row, column=col).value == "工厂":
                    return row
        return 2

    def get_column_map(ws, header_row):
        col_map = {}
        row1 = header_row
        row2 = header_row + 1
        for col in range(1, ws.max_column + 1):
            cell2 = ws.cell(row=row2, column=col) if row2 <= ws.max_row else None
            cell1 = ws.cell(row=row1, column=col)
            if cell2 and cell2.value and isinstance(cell2.value, str):
                col_name = cell2.value.replace('\n', '').strip()
            elif cell1.value and isinstance(cell1.value, str):
                col_name = cell1.value.replace('\n', '').strip()
            else:
                continue
            col_map[col] = col_name
        return col_map

    def write_data_rows(ws, data_start_row, data_rows, col_map, header_row, has_total=False):
        # 找到底部起始行（'电子版：'）
        bottom_row = None
        for row in range(data_start_row, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val and isinstance(val, str) and "电子版：" in val:
                    bottom_row = row
                    break
            if bottom_row:
                break
        if not bottom_row:
            bottom_row = ws.max_row + 1
        if bottom_row > data_start_row:
            ws.delete_rows(data_start_row, bottom_row - data_start_row)

        new_col = ws.max_column + 1
        ws.cell(row=header_row, column=new_col, value="是否符合")
        if len(data_rows) == 0:
            return

        ws.insert_rows(data_start_row, amount=len(data_rows))
        for i, row_cells in enumerate(data_rows):
            target_row = data_start_row + i
            for col_num, val in enumerate(row_cells, start=1):
                cell = ws.cell(row=target_row, column=col_num)
                if cell.__class__.__name__ == 'MergedCell':
                    continue
                if val is not None:
                    cell.value = val

        # 获取各列号
        actual_col = adj_col = diff_col = erp_col = in_col = out_col = None
        age_col_nums = []
        age_col_names = ["3个月库龄", "4-6个月库龄", "7-12个月库龄", "1-2年库龄", "2-3年库龄", "3年以上库龄", "10年以上库龄"]
        for col_num, col_name in col_map.items():
            if col_name == "实盘数量":
                actual_col = col_num
            elif col_name == "调整后数量":
                adj_col = col_num
            elif col_name == "盘盈（+）盘亏（-）数量":
                diff_col = col_num
            elif col_name == "ERP账面数量":
                erp_col = col_num
            elif col_name == "入库未记数":
                in_col = col_num
            elif col_name == "出库未记数":
                out_col = col_num
            elif col_name in age_col_names:
                age_col_nums.append(col_num)

        data_rows_count = len(data_rows) - (1 if has_total else 0)
        if data_rows_count > 0 and actual_col and adj_col and diff_col and erp_col and in_col and out_col and age_col_nums:
            age_start = min(age_col_nums)
            age_end = max(age_col_nums)
            for i in range(data_rows_count):
                target_row = data_start_row + i
                cond1 = f"N({get_column_letter(actual_col)}{target_row})=N({get_column_letter(adj_col)}{target_row})+N({get_column_letter(diff_col)}{target_row})"
                cond2 = f"N({get_column_letter(actual_col)}{target_row})=SUM({get_column_letter(age_start)}{target_row}:{get_column_letter(age_end)}{target_row})"
                cond3 = f"N({get_column_letter(adj_col)}{target_row})=N({get_column_letter(erp_col)}{target_row})+N({get_column_letter(in_col)}{target_row})-N({get_column_letter(out_col)}{target_row})"
                formula = f'=IF(AND({cond1},{cond2},{cond3}),"是","否")'
                cell = ws.cell(row=target_row, column=new_col)
                cell.value = formula
                cell.number_format = 'General'

        if has_total:
            total_row = data_start_row + len(data_rows) - 1
            ws.cell(row=total_row, column=new_col, value="")

    current_date = datetime.now()
    last_month = current_date.replace(day=1) - pd.DateOffset(days=1)
    last_month_str = last_month.strftime("%Y年%m月")

    for warehouse in warehouses:
        df_warehouse = df_merged[df_merged["仓库描述"] == warehouse].copy()
        if df_warehouse.empty:
            continue
        df_warehouse["存储位置_str"] = df_warehouse["存储位置"].astype(str).str.strip()
        df_finished = df_warehouse[~df_warehouse["存储位置_str"].str.endswith("6")].copy()
        df_gift_data = df_warehouse[df_warehouse["存储位置_str"].str.endswith("6")].copy()

        safe_warehouse = warehouse.replace("/", "_").replace("\\", "_").replace(":", "_")
        filename = f"{last_month_str}合肥美菱集团控股有限公司内（外）销产成品盘存分析表--{safe_warehouse}.xlsx"

        # 深拷贝模板
        wb = load_workbook(template_file)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        if "请输入库位" in cell.value:
                            cell.value = cell.value.replace("请输入库位", warehouse)
                        if "2026年3月" in cell.value:
                            cell.value = cell.value.replace("2026年3月", last_month_str)

        for sheet_name, data_df in [("成品", df_finished), ("赠品", df_gift_data)]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            header_row = find_header_row(ws)
            data_start_row = header_row + 2
            col_map = get_column_map(ws, header_row)

            data_rows = []
            total_qty = 0
            if not data_df.empty:
                for _, row in data_df.iterrows():
                    row_cells = [None] * ws.max_column
                    for col_num, col_name in col_map.items():
                        if col_name == "工厂":
                            val = row.get("工厂", "")
                        elif col_name == "库位":
                            val = row.get("存储位置", "")
                        elif col_name == "库位名称":
                            val = warehouse
                        elif col_name == "物料代码":
                            val = row.get("物料代码", "")
                        elif col_name == "物料描述":
                            val = row.get("物料描述", "")
                        elif col_name == "产品等级":
                            val = row.get("产品等级", "")
                        elif col_name == "单位":
                            val = row.get("单位", "")
                        elif col_name == "ERP账面数量":
                            # 修改：ERP账面数量 = 非限制使用的库存 + 冻结库存
                            unrestricted = row.get("非限制使用的库存", 0)
                            frozen = row.get("冻结库存", 0)
                            if pd.isna(unrestricted):
                                unrestricted = 0
                            if pd.isna(frozen):
                                frozen = 0
                            val = unrestricted + frozen
                            total_qty += val
                        elif col_name in ["ERP账面金额", "入库未记数", "出库未记数", "调整后数量", "实盘数量", "盘盈（+）盘亏（-）数量"]:
                            val = ""  # 留空，后期填写
                        elif col_name in age_cols:
                            val = row.get(col_name, 0)
                        else:
                            val = ""
                        row_cells[col_num-1] = val
                    data_rows.append(row_cells)
                # 合计行
                total_row = [None] * ws.max_column
                for col_num, col_name in col_map.items():
                    if col_name == "库位名称":
                        total_row[col_num-1] = "合计"
                    elif col_name == "ERP账面数量":
                        total_row[col_num-1] = total_qty
                data_rows.append(total_row)
                has_total = True
            else:
                has_total = False
            write_data_rows(ws, data_start_row, data_rows, col_map, header_row, has_total)

        # 保存到内存
        with io.BytesIO() as buf:
            wb.save(buf)
            output_files[filename] = buf.getvalue()

    # ---------- 5. 生成汇总表（成品/赠品分开） ----------
    df_merged["存储位置"] = df_merged["存储位置"].astype(str).str.strip()
    df_merged["类型"] = df_merged["存储位置"].apply(lambda x: "赠品" if x.endswith("6") else "成品")
    df_merged["账面数量"] = df_merged["非限制使用的库存"] + df_merged.get("冻结库存", 0)

    summary = df_merged.groupby(["仓库描述", "类型"])["账面数量"].sum().reset_index()
    summary_pivot = summary.pivot(index="仓库描述", columns="类型", values="账面数量").fillna(0).reset_index()
    summary_pivot.columns.name = None
    summary_pivot = summary_pivot[["仓库描述", "成品", "赠品"]]
    summary_pivot["总计"] = summary_pivot["成品"] + summary_pivot["赠品"]
    summary_pivot = summary_pivot.sort_values("总计", ascending=False)

    total_row = pd.DataFrame([["总计", summary_pivot["成品"].sum(), summary_pivot["赠品"].sum(), summary_pivot["总计"].sum()]],
                             columns=["仓库描述", "成品", "赠品", "总计"])
    summary_pivot = pd.concat([summary_pivot, total_row], ignore_index=True)

    summary_path = f"{last_month_str}库存盘点汇总表.xlsx"
    summary_bytes = io.BytesIO()
    summary_pivot.to_excel(summary_bytes, index=False)
    output_files[summary_path] = summary_bytes.getvalue()

    return output_files

# 执行生成
if st.button("🚀 生成盘存分析表"):
    if not all([ib00_file, location_file, age_file, template_file]):
        st.error("请上传所有四个文件！")
    else:
        with st.spinner("正在处理，请稍候..."):
            try:
                output_files = process_files(ib00_file, location_file, age_file, template_file)
                # 打包成 ZIP
                zip_buffer = io.BytesIO()
                with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for filename, data in output_files.items():
                        zf.writestr(filename, data)
                zip_buffer.seek(0)
                st.success("处理完成！")
                st.download_button(
                    label="📥 下载结果 (ZIP)",
                    data=zip_buffer,
                    file_name="盘存分析表结果.zip",
                    mime="application/zip"
                )
            except Exception as e:
                st.error(f"处理出错：{str(e)}")